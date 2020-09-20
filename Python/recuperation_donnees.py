""" Get the data. Just lauch it. Should not crash but if it happens, just run it again. """

from requests import *
from json import *
from math import *
from datebase import *
from local_linear_regression import *
from xlwt import Workbook
import openpyxl
import pickle
import time

def get_robuste(chaine):

    """ fonction qui refait des requetes jusqu'a avoir une reponse positive """
    
    status=503
    while status != 200:
        objet=get(chaine)
        status=objet.status_code
    return objet


def get_block_data(chaine):
    redo = True
    while(redo):
        try:
            objet=get_robuste(chaine).json()['data']
            redo = False
        except JSONDecodeError:
            pass
    return objet


def interpol_missing_day(missing_day, first_day, last_day, objet):
    if missing_day != first_day and missing_day != last_day:
        return 0.5 * (objet[date_str(missing_day-1)] + objet[date_str(missing_day+1)])
    elif missing_day == first_day:
        return objet[date_str(missing_day+1)]
    else:
        return objet[date_str(missing_day-1)]


def make_api_call_exchange_rate(indice_min, indice_max):

    api_key = '2aa8589235dcf557b78fec8085b4e02b84a3b8f2aba9e07989cb6e8f42262459'

    dt = datetime.strptime(date_str(indice_max + 1), '%Y-%m-%d')
    toTs = str(int(time.mktime(dt.timetuple())))
    limit = indice_max - indice_min 

    main_url = 'https://min-api.cryptocompare.com/data/v2/histoday?fsym=BTC&tsym=USD'
    api_call = f'{main_url}&toTs={toTs}&limit={limit}&api_key={api_key}'

    response = get(api_call)
    if response.status_code != 200:
        print('API call to cryptocompare did not work.')
    else:
        return [day['close'] for day in response.json()['Data']['Data']]
        
        
        

def update_donnees_brutes_cours():

    """ fonction qui met a jour la base donnees brutes et la base cours
    Pour des raisons de simplicite, lui donner le cours d'aujourd'hui.
    La base de donnees brutes cree contient:
    N: le nombre de blocs trouve chaque jour
    NB: le nombre de nouveaux bitcoins crees chaque jour (PAS PAR BLOC)
    frais: la somme des frais ( en BTC) recoltes par jour
    diff: la somme des difficultes par jour. Ne veut rien dire mais il suffit
    de diviser par N """

    try:
        with open('donnees_brutes','rb') as fichier:
            mon_depickler=pickle.Unpickler(fichier)
            N=mon_depickler.load()
            NB=mon_depickler.load()
            frais=mon_depickler.load()
            diff=mon_depickler.load()
            taille=mon_depickler.load()
            dernier_bloc_ok=mon_depickler.load()

        next_block=dernier_bloc_ok+1
        print('last block up to date: ',dernier_bloc_ok)
        # on retrouve le nombre de jours pour couper les listes si elles contiennent
        # plein de zeros a cause d'une erreur...
        objet=get_block_data('https://chain.api.btc.com/v3/block/'+str(dernier_bloc_ok))
        dernier_jour=floor((objet['timestamp']-1231459201)/(3600*24))
        N=N[0:dernier_jour+1]
        NB=NB[0:dernier_jour+1]
        frais=frais[0:dernier_jour+1]
        diff=diff[0:dernier_jour+1]
        taille=taille[0:dernier_jour+1]

    except FileNotFoundError:
        dernier_bloc_ok = -1
        next_block = 0
        N=[]
        NB=[]
        frais=[]
        diff=[]
        taille=[]       

        # 1231459201: 9/01/2009 00:00:01
    # on cherche le dernier bloc a recuperer
    objet = get_block_data('https://chain.api.btc.com/v3/block/latest')
    indice_max=floor((objet['timestamp']-1231459201)/(3600*24))

    last_block=objet['height']
    print('last block available: ',last_block)

    # on initialise les listes a la bonne taille

    indice_min=len(N) # ou commencer
    for t in range(indice_min,indice_max+1):
        N.append(0)
        NB.append(0)
        frais.append(0)
        diff.append(0)
        taille.append(0)
 

    # on fait des requetes a btc.com par bloc de 1000 blocs
    taille_requete=1000
    nbr_requetes_full=floor((last_block-next_block)/taille_requete)
    for j in range(0,nbr_requetes_full):
        chaine_blocs=str(taille_requete*j+next_block)
        for i in range(1,taille_requete):
            chaine_blocs=chaine_blocs+','+str(taille_requete*j+i+next_block)
        chaine='https://chain.api.btc.com/v3/block/'+chaine_blocs
        objet=get_block_data(chaine)
        for i in range(0,taille_requete):
            bloc=objet[i]
            indice=floor((bloc['timestamp']-1231459201)/(3600*24))
            N[indice]+=1
            NB[indice]+=bloc['reward_block']/100000000
            frais[indice]+=bloc['reward_fees']/100000000
            diff[indice]+=bloc['difficulty']
            taille[indice]+=bloc['size']
        dernier_bloc_ok+=taille_requete

    # on enregistre a chaque iteration comme ca crash tout le temps.   
        with open('donnees_brutes','wb') as fichier:
            mon_pickler=pickle.Pickler(fichier)
            mon_pickler.dump(N)
            mon_pickler.dump(NB)
            mon_pickler.dump(frais)
            mon_pickler.dump(diff)
            mon_pickler.dump(taille)
            mon_pickler.dump(dernier_bloc_ok)
        print('last block up to date: ',dernier_bloc_ok)
        
    # deniere requete:
    if dernier_bloc_ok < last_block:
        chaine_blocs=str(dernier_bloc_ok+1)
        i=dernier_bloc_ok+2
        while i < last_block+1:
            chaine_blocs=chaine_blocs+','+str(i)
            i+=1
        chaine='https://chain.api.btc.com/v3/block/'+chaine_blocs
        objet=get_block_data(chaine)
        if type(objet)==type(dict()):
            bloc=objet
            indice=floor((bloc['timestamp']-1231459201)/(3600*24))
            N[indice]+=1
            NB[indice]+=bloc['reward_block']/100000000
            frais[indice]+=bloc['reward_fees']/100000000
            diff[indice]+=bloc['difficulty']
            taille[indice]+=bloc['size']
            dernier_bloc_ok+=1
        else:
            for i in range(0,len(objet)):
                bloc=objet[i]
                indice=floor((bloc['timestamp']-1231459201)/(3600*24))
                N[indice]+=1
                NB[indice]+=bloc['reward_block']/100000000
                frais[indice]+=bloc['reward_fees']/100000000
                diff[indice]+=bloc['difficulty']
                taille[indice]+=bloc['size']
            dernier_bloc_ok+=len(objet)
        with open('donnees_brutes','wb') as fichier:
            mon_pickler=pickle.Pickler(fichier)
            mon_pickler.dump(N)
            mon_pickler.dump(NB)
            mon_pickler.dump(frais)
            mon_pickler.dump(diff)
            mon_pickler.dump(taille)
            mon_pickler.dump(dernier_bloc_ok)
        print('last block up to date: ',dernier_bloc_ok)


    #.................................................................
    #                    on s'occupe du cours
    #.................................................................

    # on utilise l'API de coindesk
    print('exchange rate...')

    try:
        with open('cours','rb') as fichier:
            mon_depickler=pickle.Unpickler(fichier)
            cours=mon_depickler.load()
            if len(cours) < 722:
                cours = 722*[0]

    except FileNotFoundError:
        cours = 722*[0]

    indice_min=len(cours)
    api_response = make_api_call_exchange_rate(indice_min, indice_max)
    print(api_response)
    cours += api_response

    
##    taille_requete=10
##    nbr_requetes_full=floor((indice_max-indice_min)/taille_requete)
##
##    for j in range(0,nbr_requetes_full):
##        debut=taille_requete*j+indice_min
##        fin=taille_requete*(j+1)+indice_min-1
##        chaine='http://api.coindesk.com/v1/bpi/historical/close.json?start='+\
##                date_str(debut)+'&end='+date_str(fin)
##        objet=get_robuste(chaine).json()['bpi']
##        print(objet)
##        for i in range(0,taille_requete):
##            try:
##                cours.append(objet[date_str(debut+i)])
##            except KeyError:
##                cours.append(interpol_missing_day(debut + i, debut, debut + taille_requete - 1,objet))
##        try:
##            with open('cours','wb') as fichier:
##                mon_pickler=pickle.Pickler(fichier)
##                mon_pickler.dump(cours)
##        except OSError:
##            time.sleep(1)
##            with open('cours','wb') as fichier:
##                mon_pickler=pickle.Pickler(fichier)
##                mon_pickler.dump(cours)
##        print('last day up to date: ',date_str(fin))
##
##
##    # derniere requete
##    debut=taille_requete*nbr_requetes_full+indice_min
##    if debut <= indice_max-1:
##        chaine='http://api.coindesk.com/v1/bpi/historical/close.json?start='+\
##                date_str(debut)+'&end='+date_str(indice_max-1)
##        objet=get_robuste(chaine).json()['bpi']
##        for i in range(indice_min+taille_requete*nbr_requetes_full,indice_max):
##            cours.append(objet[date_str(i)])
    with open('cours','wb') as fichier:
            mon_pickler=pickle.Pickler(fichier)
            mon_pickler.dump(cours)
##    print('last day up to date: ',date_str(indice_max-1))

def  creation_base_donnees_traites():

    """ cree une base de donnees utilisable a partir de la base de donnees
    brutes """

    with open('donnees_brutes','rb') as fichier:
        mon_depickler=pickle.Unpickler(fichier)
        N=mon_depickler.load()
        NB=mon_depickler.load()
        frais=mon_depickler.load()
        diff=mon_depickler.load()
        taille=mon_depickler.load()
        
    # on laisse de cote le dernier jour, mal estime.
    # Inutile pour le cours: deja fait.
    N=N[:-1]
    NB=NB[:-1]
    frais=frais[:-1]
    diff=diff[:-1]
    taille=taille[:-1]

    with open('cours','rb') as fichier2:
        mon_depickler2=pickle.Unpickler(fichier2)
        cours=mon_depickler2.load()

    K=3600*24*10**12

    # on commence par faire des moyennes la ou on a juste fait la somme
    diff_m=[d/n for d,n in zip(diff,N)] # difficulte par bloc
    NB_m=[nb/n for nb,n in zip(NB,N)] # nouveaux bitcoins crees par bloc
    frais_m=[f/n for f,n in zip(frais,N)] # frais par bloc
    S=[t/n for t,n in zip(taille,N)] # taille des blocs en bytes

    proba=[1/(a*2**32) for a in diff_m] # proba de trouver un bloc en un hash
    R=[(nb+f)*c for nb,f,c in zip(NB_m,frais_m,cours)]
    # recompense par bloc en dollars
    P=[K*a*b for a,b in zip(R,proba)] # payoff

    Q_hat=[n/(K*p) for n,p in zip(N,proba)]
    fenetre=15 
    retour_arriere=10*fenetre # de combien revient on en arriere dans la
    # regression lineaire locale
    
    
    print('local linear regression...')
    try:
        with open('donnees','rb') as fifi:
            mon_dede = pickle.Unpickler(fifi)
            R_inutile=mon_dede.load()
            P_inutile=mon_dede.load()
            Q=mon_dede.load()
    except FileNotFoundError:
        Q = 150*[0]
        
    old_len=len(Q)
    Q_smooth=local_linear_regression(\
        Q_hat[old_len-retour_arriere:],range(old_len-retour_arriere,len(Q_hat))\
        ,fenetre)
    Q=Q[:old_len-int(retour_arriere/2)]+Q_smooth[int(retour_arriere/2):]

    with open('donnees','wb') as fichier:
        mon_pickler = pickle.Pickler(fichier)
        mon_pickler.dump(R)
        mon_pickler.dump(P)
        mon_pickler.dump(Q)
        mon_pickler.dump(cours)
        mon_pickler.dump(N)
        mon_pickler.dump(S)

def export_excel():
    with open('donnees','rb') as fichier:
        mon_depickler=pickle.Unpickler(fichier)
        R=mon_depickler.load()
        P=mon_depickler.load()
        Q=mon_depickler.load()
        cours=mon_depickler.load()
        N=mon_depickler.load()
        S=mon_depickler.load()

    with open('donnees_brutes','rb') as fichier2:
        mon_depickler2=pickle.Unpickler(fichier2)
        N=mon_depickler2.load()
        NB=mon_depickler2.load()
        frais=mon_depickler2.load()
        diff=mon_depickler2.load()
        taille=mon_depickler2.load()

    book=Workbook()
    feuille1 = book.add_sheet('feuille 1')
    for t in range(1,len(R)):
        # drop the first date (which anyway we do not care about)
        # so indices are the same on Matlab (where indices start at 1) and Python
        # (where indices start at 0)
        feuille1.write(t-1,0,R[t])
        feuille1.write(t-1,1,P[t])
        feuille1.write(t-1,2,Q[t])
        feuille1.write(t-1,3,cours[t])
        feuille1.write(t-1,4,N[t])
        feuille1.write(t-1,5,S[t])
        feuille1.write(t-1,6,NB[t])
        feuille1.write(t-1,7,frais[t])
        feuille1.write(t-1,8,diff[t])
    book.save('database.xls')

def get_exchange_rate_kaiko():
    wb=openpyxl.load_workbook('kaiko-daily-price.xlsx')
    sheet = wb.get_sheet_by_name('kaiko-daily-price')
    kaiko = 554*[0]
    for i in range(2, sheet.max_row + 1):
        kaiko.append(float(sheet.cell(row=i,column=2).value))
    return kaiko

def charger_donnees():
    with open('donnees', 'rb') as fichier:
        mon_depickler = pickle.Unpickler(fichier)
        R=mon_depickler.load()
        P=mon_depickler.load()
        Q=mon_depickler.load()
        cours=mon_depickler.load()
    return R,P,Q,cours
    

        
if __name__ == "__main__":
    update_donnees_brutes_cours()
    creation_base_donnees_traites()
    export_excel()




