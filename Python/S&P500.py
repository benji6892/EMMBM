""" For our two periods, estimates the parameters mu and sigma2 corresponding
to the S&P 500 series""".

import openpyxl
from estimation import estimation_alpha_sigma2

# first period

wb=openpyxl.load_workbook('S&P1.xlsx')
sheet = wb.get_sheet_by_name('Feuille1')
maxi=sheet.max_row

price=[]
for i in range(2,maxi+1):
    price.append(float(sheet.cell(row=i,column=6).value))

alpha,sigma2=estimation_alpha_sigma2(price)
print('\npremiere periode: \nmu: ',365*(alpha-0.5*sigma2),'\nsigma2: ',365*sigma2)

# dsecond period

wb=openpyxl.load_workbook('S&P2.xlsx')
sheet = wb.get_sheet_by_name('Feuille1')
maxi=sheet.max_row

price=[]
for i in range(2,maxi+1):
    price.append(float(sheet.cell(row=i,column=6).value))

alpha,sigma2=estimation_alpha_sigma2(price)
print('\ndeuxieme periode: \nmu: ',365*(alpha-0.5*sigma2),'\nsigma2: ',365*sigma2)
