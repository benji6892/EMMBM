""" ce module prend en input la barriere et le Q simule dans matlab (en tenant
    compte du halving et il fait les graphiques pour tous les modeles. """


from matplotlib.pyplot import *
from datebase import *
import openpyxl
import pickle
from numpy import *
from simulation_Q import *

with open('donnees', 'rb') as fichier:
    mon_depickler = pickle.Unpickler(fichier)
    R=mon_depickler.load()
    P=mon_depickler.load()
    Q=mon_depickler.load()

lw=3

#*******************************************************************************
#                             premiere periode
#*******************************************************************************


### on importe les donnees de matlab
##wb=openpyxl.load_workbook('periode1.xlsx')
##sheet = wb.get_sheet_by_name('Feuil1')
##maxi=sheet.max_row
##
##Q_sim=[]
##Bt=[]
##for i in range(1,maxi+1):
##    Q_sim.append(sheet.cell(row=i,column=1).value)
##    Bt.append(sheet.cell(row=i,column=2).value)
##    
##d=722
##f=1483
##h=1419
##
### On calcule le hashrate dans le modèle avec electricite
##Rp=R[d:f+1]
##Q0=Q[d-1]
##a0=0.0034209
##B_tilde0=33504
##C0=680
##x0=estimation_a_B_tilde(a0,B_tilde0,C0,R,Q,d,f,simulationQ_cv)
##a=x0['x'][0]
##B_tilde=x0['x'][1]
##Q_sim_C,P_sim_C=simulationQ_cv(a,B_tilde,Rp,Q0,C0)
##Q_sim_base,P_sim_base=simulationQ(a0,B_tilde0,Rp,Q0,1)
##
##xdates=[date_base(j) for j in range(d,f+1)]
##
##figure(1)
##plot(xdates,log(Q_sim),'b',linewidth=lw,\
##     label='log simulated hashrate with halving')
##plot(xdates,log(Q_sim_C),'c',linewidth=lw,\
##     label='log simulated hashrate with electricity')
##plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log observed hashrate')
##plot(xdates,log(Q_sim_base),'g--',linewidth=lw,\
##     label='log simulated hashrate base model')
###axvline(x=date_base(h),color='k',linestyle='dashed',\
###        linewidth=lw,label='halving')
##gcf().autofmt_xdate()
##legend()
##show()
##
##T=len(Q_sim)
##At=exp(a*arange(0,T))
##P_sim=[r/q for r,q in zip(Rp,Q_sim)]
##figure(2)
##plot(xdates,[144*a*p for a,p in zip(At,P_sim)],'b',linewidth=lw,\
##     label='simulated detrended payoff with halving')
##plot(xdates,[a*p for a,p in zip(At,P[d:f+1])],'r',linewidth=lw,\
##     label='observed payoff')
##plot(xdates,[a*p for a,p in zip(At,P_sim_base)],'g--',linewidth=lw,\
##     label='simulated detrended payoff base model')
##plot(xdates,[a*b for a,b in zip(At,Bt)],'k--',linewidth=lw,\
##     label='barrier with halving')
##gcf().autofmt_xdate()
##legend()
##show()



#*******************************************************************************
#                             deuxieme periode
#*******************************************************************************


# on importe les donnees de matlab
wb=openpyxl.load_workbook('periode2.xlsx')
sheet = wb.get_sheet_by_name('Feuil1')
maxi=sheet.max_row

Q_sim=[]
Bt=[]
for i in range(1,maxi+1):
    Q_sim.append(sheet.cell(row=i,column=1).value)
    Bt.append(sheet.cell(row=i,column=2).value)
    
d=2091
f=3003
h=2738

# on recalcule le hashrate dans le modèle de base
Rp=R[d:f+1]
Q0=Q[d-1]
a0=0.002391
B_tilde0=5.27
C0=0.84
x0=estimation_a_B_tilde(a0,B_tilde0,C0,R,Q,d,f,simulationQ_cv)
a=x0['x'][0]
B_tilde=x0['x'][1]
Q_sim_C,P_sim_C=simulationQ_cv(a,B_tilde,Rp,Q0,C0)
Q_sim_base,P_sim_base=simulationQ(a0,B_tilde0,Rp,Q0,1)

xdates=[date_base(j) for j in range(d,f+1)]

figure(1)
plot(xdates,log(Q_sim),'b',linewidth=lw,\
     label='log simulated hashrate with halving')
plot(xdates,log(Q_sim_C),'r',linewidth=lw,\
     label='log simulated hashrate with electricity')
plot(xdates,log(Q[d:f+1]),'k',linewidth=lw,label='log observed hashrate')
plot(xdates,log(Q_sim_base),'g--',linewidth=lw,\
     label='log simulated hashrate base model')
#axvline(x=date_base(h),color='k',linestyle='dashed',\
#        linewidth=lw,label='halving')
gcf().autofmt_xdate()
legend()
show()

T=len(Q_sim)
At=exp(a0*arange(0,T))
P_sim=[r/q for r,q in zip(Rp,Q_sim)]
figure(2)
plot(xdates,[144*a*p for a,p in zip(At,P_sim)],'b',linewidth=lw,\
     label='simulated detrended payoff with halving')
plot(xdates,[a*p for a,p in zip(At,P[d:f+1])],'r',linewidth=lw,\
     label='observed payoff')
plot(xdates,[a*p for a,p in zip(At,P_sim_base)],'g--',linewidth=lw,\
     label='simulated detrended payoff base model')
plot(xdates,[a*b for a,b in zip(At,Bt)],'k--',linewidth=lw,\
     label='barrier with halving')
gcf().autofmt_xdate()
legend()
show()
