""" Plots graphs for the model with halving. Needs as input an excel file with
the barrier computed on Matlab. """

import csv
import openpyxl
from recuperation_donnees import charger_donnees
from simulation_Q_base import *

R,P,Q,cours=charger_donnees()

lw=3
fg=(12,6)
ft='xx-large'

###*******************************************************************************
###                             premiere periode
###*******************************************************************************
##
##
### on importe les donnees de matlab
##wb=openpyxl.load_workbook('periode1.xlsx')
##sheet = wb.get_sheet_by_name('Feuil1')
##maxi=sheet.max_row
##
##Q_sim=[]
##Bt=[] # barrier
##for i in range(1,maxi+1):
##    Q_sim.append(sheet.cell(row=i,column=1).value)
##    Bt.append(sheet.cell(row=i,column=2).value)
##
##d=812
##f=1483
##h=1419
##
### with this (no flat) barrier, we now recompute the hashrate using the baseline
### model.
##Rp=R[d:f+1]
##Q0=Q[d-1]
##a=0.0032334
##B_tilde=25996
##Q_sim_base,P_sim_base=simulationQ(a,B_tilde,Rp,Q0)
##
##xdates=[date_base(j) for j in range(d,f+1)]
##
##fig, ax=subplots(figsize=fg)
##plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log Q')
##ax.plot(xdates,log(Q_sim),'b--',linewidth=lw,\
##     label=r'log $Q^{sim}$ halving')
##plot(xdates,log(Q_sim_base),'g:',linewidth=lw,\
##     label='log $Q^{sim}$ baseline')
##axvline(x=date_base(h),color='k',linestyle='-.',\
##        linewidth=lw,label='halving')
##gcf().autofmt_xdate()
##ax.tick_params(axis='both',labelsize=ft)
##ax.xaxis.set_major_locator(MonthLocator(interval=9))
##legend(fontsize=ft)
##show()
##
##P_sim = [144*r/q for r,q in zip(Rp,Q_sim)]
##Pp = P[d:f+1]
##
##fig2, ax2 = subplots(figsize=fg)
##ax2.plot(xdates,P_sim,'b',linewidth=lw,\
##         linestyle='--',label='simulated payoff')
##ax2.plot(xdates,Pp,'g',linewidth=lw,\
##         label='observed payoff')
##ax2.plot(xdates,Bt,'r',linewidth=lw,linestyle='-.',\
##         label='barrier')
##if (d<1419) and (f>1419):
##    ax2.axvline(x=date_base(1419),color='k',linestyle=':',\
##                linewidth=lw,label='halving')
##if (d<2738) and (f>2738):
##    ax2.axvline(x=date_base(2738),color='k',linestyle=':',\
##                linewidth=lw,label='halving')  
##gcf().autofmt_xdate()
##ax2.tick_params(axis='both',labelsize=ft)
##ax2.xaxis.set_major_locator(MonthLocator(interval=9))
##
##legend(fontsize=ft)
##gcf().subplots_adjust(left=0.15)
##ax2.margins(0.03)
##
##show()

###*******************************************************************************
###                             deuxieme periode
###*******************************************************************************
##
##
### on importe les donnees de matlab
##wb=openpyxl.load_workbook('periode2.xlsx')
##sheet = wb.get_sheet_by_name('Feuil1')
##maxi=sheet.max_row
##
##Q_sim=[]
##Bt=[]
##for i in range(1,maxi+1):
##    Q_sim.append(sheet.cell(row=i,column=1).value)
##    Bt.append(sheet.cell(row=i,column=2).value)
##
##d=2091
##f=3003
##h=2738
##
### on recalcule le hashrate dans le modèle de base
##Rp=R[d:f+1]
##Q0=Q[d-1]
##a=0.00207
##B_tilde=5.30
##Q_sim_base,P_sim_base=simulationQ(a,B_tilde,Rp,Q0)
##
##xdates=[date_base(j) for j in range(d,f+1)]
##
##fig, ax=subplots(figsize=fg)
##plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log Q')
##ax.plot(xdates,log(Q_sim),'b--',linewidth=lw,\
##     label=r'log $Q^{sim}$ halving')
##plot(xdates,log(Q_sim_base),'g:',linewidth=lw,\
##     label='log $Q^{sim}$ baseline')
##axvline(x=date_base(h),color='k',linestyle='-.',\
##        linewidth=lw,label='halving')
##gcf().autofmt_xdate()
##ax.tick_params(axis='both',labelsize=ft)
##ax.xaxis.set_major_locator(MonthLocator(interval=9))
##legend(fontsize=ft)
##show()
##
##P_sim = [144*r/q for r,q in zip(Rp,Q_sim)]
##Pp = P[d:f+1]
##
##fig2, ax2 = subplots(figsize=fg)
##ax2.plot(xdates,P_sim,'b',linewidth=lw,\
##         linestyle='--',label='simulated payoff')
##ax2.plot(xdates,Pp,'g',linewidth=lw,\
##         label='observed payoff')
##ax2.plot(xdates,Bt,'r',linewidth=lw,linestyle='-.',\
##         label='barrier')
##if (d<1419) and (f>1419):
##    ax2.axvline(x=date_base(1419),color='k',linestyle=':',\
##                linewidth=lw,label='halving')
##if (d<2738) and (f>2738):
##    ax2.axvline(x=date_base(2738),color='k',linestyle=':',\
##                linewidth=lw,label='halving')  
##gcf().autofmt_xdate()
##ax2.tick_params(axis='both',labelsize=ft)
##ax2.xaxis.set_major_locator(MonthLocator(interval=9))
##
##legend(fontsize=ft)
##gcf().subplots_adjust(left=0.15)
##ax2.margins(0.03)
##
##show()
##
### Et maintenant un graphique qui montre juste les deux barrieres.
##
##fig,ax=subplots(figsize=(11,6))
##axvline(x=date_base(h),color='k',linestyle=':',\
##        linewidth=lw,label='halving')
##ax.plot(xdates,[Bt[0]*exp(-0.00232*t) for t in range(0,f+1-d)],'r',linewidth=lw,\
##        linestyle='--',label='barrier without halving')
##ax.plot(xdates,Bt,'b',linewidth=lw,\
##     label='barrier with halving')
##gcf().autofmt_xdate()
##ax.tick_params(axis='both',labelsize=ft)
##ax.xaxis.set_major_locator(MonthLocator(interval=9))
##legend(fontsize=ft)
##show()


#*******************************************************************************
#                             third period
#*******************************************************************************

Q_sim=[]
Bt = []
# on importe les donnees de matlab
with open('periode3.csv') as csvfile:
    file = csv.reader(csvfile, delimiter=';', quotechar='"')
    for row in file:
        Q_sim.append(double(row[0].replace(',', '.')))
        Bt.append(double(row[1].replace(',', '.')))

d=3491
f=4271
h=4140

# on recalcule le hashrate dans le modèle de base
Rp=R[d:f+1]
Q0=Q[d-1]
a=0.00209
B_tilde=0.6
Q_sim_base,P_sim_base=simulationQ(a,B_tilde,Rp,Q0)

xdates=[date_base(j) for j in range(d,f+1)]

fig, ax=subplots(figsize=fg)
plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log Q')
ax.plot(xdates,log(Q_sim),'b--',linewidth=lw,\
     label=r'log $Q^{sim}$ halving')
plot(xdates,log(Q_sim_base),'g:',linewidth=lw,\
     label='log $Q^{sim}$ baseline')
axvline(x=date_base(h),color='k',linestyle='-.',\
        linewidth=lw,label='halving')
gcf().autofmt_xdate()
ax.tick_params(axis='both',labelsize=ft)
ax.xaxis.set_major_locator(MonthLocator(interval=9))
legend(fontsize=ft)
show()

P_sim = [144*r/q for r,q in zip(Rp,Q_sim)]
Pp = P[d:f+1]

fig2, ax2 = subplots(figsize=fg)
ax2.plot(xdates,P_sim,'b',linewidth=lw,\
         linestyle='--',label='simulated payoff')
ax2.plot(xdates,Pp,'g',linewidth=lw,\
         label='observed payoff')
ax2.plot(xdates,Bt,'r',linewidth=lw,linestyle='-.',\
         label='barrier')
if (d<1419) and (f>1419):
    ax2.axvline(x=date_base(1419),color='k',linestyle=':',\
                linewidth=lw,label='halving')
if (d<2738) and (f>2738):
    ax2.axvline(x=date_base(2738),color='k',linestyle=':',\
                linewidth=lw,label='halving')  
gcf().autofmt_xdate()
ax2.tick_params(axis='both',labelsize=ft)
ax2.xaxis.set_major_locator(MonthLocator(interval=9))

legend(fontsize=ft)
gcf().subplots_adjust(left=0.15)
ax2.margins(0.03)

show()

# Et maintenant un graphique qui montre juste les deux barrieres.

fig,ax=subplots(figsize=(11,6))
axvline(x=date_base(h),color='k',linestyle=':',\
        linewidth=lw,label='halving')
ax.plot(xdates,[Bt[0]*exp(-0.00232*t) for t in range(0,f+1-d)],'r',linewidth=lw,\
        linestyle='--',label='barrier without halving')
ax.plot(xdates,Bt,'b',linewidth=lw,\
     label='barrier with halving')
gcf().autofmt_xdate()
ax.tick_params(axis='both',labelsize=ft)
ax.xaxis.set_major_locator(MonthLocator(interval=9))
legend(fontsize=ft)
show()
