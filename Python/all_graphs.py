""" Plots graphs for the model with halving. Needs as input an excel file with
the barrier computed on Matlab. """

import openpyxl
from recuperation_donnees import charger_donnees
from simulation_Q_base import *
from simulation_Q_delay import *

R,P,Q,cours=charger_donnees()

lw=3
fg=(7.5,6)
ft='xx-large'

#*******************************************************************************
#                             premiere periode
#*******************************************************************************


# on importe les donnees de matlab
wb=openpyxl.load_workbook('periode1.xlsx')
sheet = wb.get_sheet_by_name('Feuil1')
maxi=sheet.max_row

Q_sim_hf=[]
Bt_hf=[] # barrier finite horizon
for i in range(1,maxi+1):
    Q_sim_hf.append(sheet.cell(row=i,column=1).value)
    Bt_hf.append(sheet.cell(row=i,column=2).value)

d=812
f=1483
h=1419

# with this (no flat) barrier, we now recompute the hashrate using the baseline
# model.
Rp=R[d:f+1]
Q0=Q[d-1]
a=0.0032334
B_tilde=25996
Q_sim_base,P_sim_base=simulationQ(a,B_tilde,Rp,Q0)

# now model with delay
a_delay = 0.00300
B_tilde_delay = 21963
delta = 11.5
C_sim_delay,Q_sim_delay,P_sim_delay=simulationQdelay(a_delay,\
                                                     B_tilde_delay,delta,Rp,Q0)

xdates=[date_base(j) for j in range(d,f+1)]

fig, ax=subplots(figsize=fg)
plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log Q')
ax.plot(xdates,log(Q_sim_hf),'b--',linewidth=lw,\
     label=r'log $Q^{sim}$ halving')
ax.plot(xdates,log(Q_sim_delay),'m-.',linewidth=lw,\
     label=r'log $Q^{sim}$ delay')
plot(xdates,log(Q_sim_base),'g:',linewidth=lw,\
     label='log $Q^{sim}$ baseline')
axvline(x=date_base(h),color='k',linestyle='-.',\
        linewidth=lw,label='halving')
gcf().autofmt_xdate()
ax.tick_params(axis='both',labelsize=ft)
ax.xaxis.set_major_locator(MonthLocator(interval=9))
legend(fontsize=ft)
show()


#*******************************************************************************
#                             deuxieme periode
#*******************************************************************************


# on importe les donnees de matlab
wb=openpyxl.load_workbook('periode2.xlsx')
sheet = wb.get_sheet_by_name('Feuil1')
maxi=sheet.max_row

Q_sim_hf=[]
Bt_hf=[]
for i in range(1,maxi+1):
    Q_sim_hf.append(sheet.cell(row=i,column=1).value)
    Bt_hf.append(sheet.cell(row=i,column=2).value)

d=2091
f=3003
h=2738

# on recalcule le hashrate dans le modèle de base
Rp=R[d:f+1]
Q0=Q[d-1]
a=0.00207
B_tilde=5.30
Q_sim_base,P_sim_base=simulationQ(a,B_tilde,Rp,Q0)

# now model with delay
a_delay = 0.002475
B_tilde_delay = 5.100
delta = 46.5
C_sim_delay,Q_sim_delay,P_sim_delay=simulationQdelay(a_delay,\
                                                     B_tilde_delay,delta,Rp,Q0)

xdates=[date_base(j) for j in range(d,f+1)]

fig, ax=subplots(figsize=fg)
plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log Q')
ax.plot(xdates,log(Q_sim_hf),'b--',linewidth=lw,\
     label=r'log $Q^{sim}$ halving')
ax.plot(xdates,log(Q_sim_delay),'m-.',linewidth=lw,\
     label=r'log $Q^{sim}$ delay')
plot(xdates,log(Q_sim_base),'g:',linewidth=lw,\
     label='log $Q^{sim}$ baseline')
axvline(x=date_base(h),color='k',linestyle='-.',\
        linewidth=lw,label='halving')
gcf().autofmt_xdate()
ax.tick_params(axis='both',labelsize=ft)
ax.xaxis.set_major_locator(MonthLocator(interval=9))
#legend(fontsize=ft)
show()


#*******************************************************************************
#                             third period
#*******************************************************************************


# on importe les donnees de matlab
wb=openpyxl.load_workbook('periode2.xlsx')
sheet = wb.get_sheet_by_name('Feuil1')
maxi=sheet.max_row

Q_sim_hf=[]
Bt_hf=[]
for i in range(1,maxi+1):
    Q_sim_hf.append(sheet.cell(row=i,column=1).value)
    Bt_hf.append(sheet.cell(row=i,column=2).value)

d=3491
f=4271
h=4140

# on recalcule le hashrate dans le modèle de base
Rp=R[d:f+1]
Q0=Q[d-1]
a=0.00207
B_tilde=5.30
Q_sim_base,P_sim_base=simulationQ(a,B_tilde,Rp,Q0)

# now model with delay
a_delay = 0.002475
B_tilde_delay = 0.6
delta = 46.5
C_sim_delay,Q_sim_delay,P_sim_delay=simulationQdelay(a_delay,\
                                                     B_tilde_delay,delta,Rp,Q0)

xdates=[date_base(j) for j in range(d,f+1)]

fig, ax=subplots(figsize=fg)
plot(xdates,log(Q[d:f+1]),'r',linewidth=lw,label='log Q')
ax.plot(xdates,log(Q_sim_hf),'b--',linewidth=lw,\
     label=r'log $Q^{sim}$ halving')
ax.plot(xdates,log(Q_sim_delay),'m-.',linewidth=lw,\
     label=r'log $Q^{sim}$ delay')
plot(xdates,log(Q_sim_base),'g:',linewidth=lw,\
     label='log $Q^{sim}$ baseline')
axvline(x=date_base(h),color='k',linestyle='-.',\
        linewidth=lw,label='halving')
gcf().autofmt_xdate()
ax.tick_params(axis='both',labelsize=ft)
ax.xaxis.set_major_locator(MonthLocator(interval=9))
#legend(fontsize=ft)
show()
